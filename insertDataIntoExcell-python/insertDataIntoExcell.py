import openpyxl

def getLastSerialNumber(sheet_obj):
    return sheet_obj.max_row

def insertPatientRecord(patient_sheet_obj, PatientID, firstName, lastName, condition):
    
    lastSerialNumber = getLastSerialNumber(patient_sheet_obj)
    
    patient_sheet_obj['A'+str(lastSerialNumber+1)] = lastSerialNumber
    patient_sheet_obj['B'+str(lastSerialNumber+1)] = PatientID
    patient_sheet_obj['C'+str(lastSerialNumber+1)] = firstName
    patient_sheet_obj['D'+str(lastSerialNumber+1)] = lastName
    patient_sheet_obj['E'+str(lastSerialNumber+1)] = condition
    
def insertMedicationRecord(PatientID, medication):
    
    medication_sheet_obj = wb_obj['Medications']
    lastSerialNumber = getLastSerialNumber(medication_sheet_obj)
    
    medication_sheet_obj['A'+str(lastSerialNumber+1)] = lastSerialNumber
    medication_sheet_obj['B'+str(lastSerialNumber+1)] = PatientID
    medication_sheet_obj['C'+str(lastSerialNumber+1)] = medication

def insertActivityRecord(PatientID, activity):
    
    activity_sheet_obj = wb_obj['Activities']
    lastSerialNumber = getLastSerialNumber(activity_sheet_obj)
    
    activity_sheet_obj['A'+str(lastSerialNumber+1)] = lastSerialNumber
    activity_sheet_obj['B'+str(lastSerialNumber+1)] = PatientID
    activity_sheet_obj['C'+str(lastSerialNumber+1)] = activity
    
path = "./sample.xlsx"

wb_obj = openpyxl.load_workbook(path)
patient_sheet_obj = wb_obj['Patient Records']
print(patient_sheet_obj)

lastSerialNumber = patient_sheet_obj.max_row
running = True

while running:
    PatientID = 1 or patient_sheet_obj['B'+str(lastSerialNumber)].value + 1
    firstName = input('Enter patient first name: ')
    lastName = input('Enter patient last name: ')
    condition = input('Enter patient health problem: ')
    medication = input('Enter all medication: ').split(' ')
    activities = input('Enter all activities: ').split(' ')
    
    insertPatientRecord(patient_sheet_obj, PatientID, firstName, lastName, condition)
    for meds in medication:
        insertMedicationRecord(PatientID, meds)
    for activity in activities:
        insertActivityRecord(PatientID, activity)
        
    wb_obj.save(path)
    # patient_sheet_obj['A'+str(lastSerialNumber+1)] = lastSerialNumber
    # patient_sheet_obj['B'+str(lastSerialNumber+1)] = PatientID
    # patient_sheet_obj['C'+str(lastSerialNumber+1)] = firstName
    # patient_sheet_obj['D'+str(lastSerialNumber+1)] = lastName
    # patient_sheet_obj['E'+str(lastSerialNumber+1)] = condition
    # patient_sheet_obj['F'+str(lastSerialNumber+1)] = medication
    # patient_sheet_obj['G'+str(lastSerialNumber+1)] = activities
    cont = input('Insert new patient record (yes/no): ')
    if cont != 'yes':
        running = False


# Patient Name :
# Max
# Mustermann
# Gesundheitsproblem: Herzinsuffizienz              Health problem: heart failure
# Medikamente:                                      Medicines
# Blutverdünner Thomapyrin                          blood-thinner thomapyrin Aspirin
# 
# Tätigkeiten                                       activities
# Blutdruck                                         blood-pressure Pulse-rate blood-sugar-levels Blood-drawn
# Pulsschlag                                        Pulse-rate
# Blutzucker spiegel                                blood-sugar-levels
# Blut abgenommen                                   Blood-drawn
